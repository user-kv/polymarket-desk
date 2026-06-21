"""
lib/report.py
Refreshes the live dashboard's data snapshot + generates tracker.xlsx from the ledger.

Dashboard: there is now ONE dashboard — desk/dashboard.html (live, reads
desk/dashboard_state.json). The old standalone papertrader/dashboard.html was retired.
`generate_dashboard()` now just refreshes that single snapshot via desk/export_state.

tracker.xlsx includes:
  - Weather tab: auto-synced from bets.csv
  - Sports tab: manual template with de-vig + CLV formula columns
"""

import os
import sys
import json
import csv
import logging
import subprocess
from datetime import datetime, timezone

logger = logging.getLogger("report")

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
REPO_ROOT = os.path.dirname(BASE_DIR)
DATA_DIR = os.path.join(BASE_DIR, "data")
BETS_PATH = os.path.join(DATA_DIR, "bets.csv")
BANKROLL_PATH = os.path.join(DATA_DIR, "bankroll.json")
DESK_DASHBOARD = os.path.join(REPO_ROOT, "desk", "dashboard.html")
TRACKER_PATH = os.path.join(BASE_DIR, "tracker.xlsx")


def _load_bets():
    if not os.path.exists(BETS_PATH):
        return []
    with open(BETS_PATH, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _load_bankroll():
    if not os.path.exists(BANKROLL_PATH):
        return {"balance": 500.0, "start": 500.0, "history": []}
    with open(BANKROLL_PATH) as f:
        return json.load(f)


def _real_bets(bets):
    """Exclude TEST bets from all statistics — they stay visible in tables but never count."""
    return [b for b in bets if b.get("is_test", "N") != "Y"]


def _pnl_stats(bets):
    bets = _real_bets(bets)
    settled = [b for b in bets if b.get("status") == "settled"]
    won = [b for b in settled if b.get("result") == "WON"]
    total_pnl = sum(float(b.get("pnl", 0)) for b in settled)
    win_rate = len(won) / len(settled) if settled else 0.0
    avg_edge = (
        sum(float(b.get("edge_pct", 0)) for b in bets) / len(bets) if bets else 0.0
    )
    return {
        "n_total": len(bets),
        "n_settled": len(settled),
        "n_won": len(won),
        "n_lost": len(settled) - len(won),
        "n_open": len([b for b in bets if b.get("status") == "open"]),
        "total_pnl": round(total_pnl, 2),
        "win_rate": round(win_rate * 100, 1),
        "avg_edge": round(avg_edge, 1),
    }


def _calibration_rows(bets):
    """Group settled bets by 10pt probability buckets; compare model prob vs actual win rate."""
    buckets = {}  # key: (lo, hi) -> [won, total]
    for b in _real_bets(bets):
        if b.get("status") != "settled":
            continue
        prob = float(b.get("ensemble_prob", 0))
        bucket_key = int(prob * 10) * 10  # 0,10,20,...,90
        lo, hi = bucket_key, bucket_key + 10
        entry = buckets.setdefault((lo, hi), [0, 0])
        entry[1] += 1
        if b.get("result") == "WON":
            entry[0] += 1
    rows = []
    for (lo, hi), (won, total) in sorted(buckets.items()):
        rows.append({
            "prob_range": f"{lo}%-{hi}%",
            "n_bets": total,
            "actual_win_pct": round(won / total * 100, 1) if total else 0,
        })
    return rows


def generate_dashboard():
    """Refresh the single live dashboard's data snapshot (desk/dashboard_state.json,
    which desk/dashboard.html reads). Returns the dashboard path.

    The standalone papertrader/dashboard.html was retired in favour of one live
    dashboard; this shim keeps every existing call site (post-scan, post-settle,
    `report`) refreshing that single source of truth. Best-effort: a failed refresh
    never breaks scan/settle."""
    try:
        subprocess.run(
            [sys.executable, "-m", "desk.export_state"],
            cwd=REPO_ROOT, check=False,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        logger.info("Dashboard state refreshed (desk/dashboard_state.json).")
    except Exception as e:
        logger.warning(f"Dashboard state refresh failed: {e}")
    return DESK_DASHBOARD


def generate_tracker():
    """Write tracker.xlsx with Weather tab (from bets.csv) and Sports tab (template)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return None

    bets = _load_bets()
    br = _load_bankroll()
    wb = openpyxl.Workbook()

    # ---- WEATHER TAB ----
    ws = wb.active
    ws.title = "Weather Bets"

    hdr_fill = PatternFill("solid", fgColor="2563EB")
    hdr_font = Font(bold=True, color="FFFFFF")

    weather_cols = [
        "Date", "City", "Market Question", "Bucket (°F)",
        "Edge (pt)", "Model Prob", "Ask Price", "Stake",
        "Shares", "Status", "Result", "Actual High (°F)", "P&L ($)",
        "GFS Mean (°F)", "ECMWF Mean (°F)", "N Members", "Is Test",
    ]

    for col, hdr in enumerate(weather_cols, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_n, b in enumerate(bets, 2):
        low_f = b.get("bucket_low_f", "")
        high_f = b.get("bucket_high_f", "")
        bucket_str = f"{low_f}–{high_f}°F"

        vals = [
            b.get("timestamp", "")[:10],
            b.get("city", ""),
            b.get("question", ""),
            bucket_str,
            b.get("edge_pct", ""),
            b.get("ensemble_prob", ""),
            b.get("ask_price", ""),
            b.get("stake", ""),
            b.get("shares", ""),
            b.get("status", ""),
            b.get("result", ""),
            b.get("actual_high_f", ""),
            b.get("pnl", ""),
            b.get("gfs_mean_f", ""),
            b.get("ecmwf_mean_f", ""),
            b.get("n_members", ""),
            b.get("is_test", "N"),
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row_n, column=col, value=val)

    # Colour WON/LOST rows
    won_fill = PatternFill("solid", fgColor="D4EDDA")
    lost_fill = PatternFill("solid", fgColor="F8D7DA")
    for row_n, b in enumerate(bets, 2):
        if b.get("result") == "WON":
            for col in range(1, len(weather_cols) + 1):
                ws.cell(row=row_n, column=col).fill = won_fill
        elif b.get("result") == "LOST":
            for col in range(1, len(weather_cols) + 1):
                ws.cell(row=row_n, column=col).fill = lost_fill

    # Column widths
    widths = [12, 10, 55, 16, 8, 10, 10, 7, 8, 9, 7, 14, 8, 12, 14, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # Summary area top-right
    summary_col = len(weather_cols) + 2
    ws.cell(row=1, column=summary_col, value="Summary").font = Font(bold=True)
    ws.cell(row=2, column=summary_col, value="Bankroll Start")
    ws.cell(row=2, column=summary_col + 1, value=br.get("start", 500))
    ws.cell(row=3, column=summary_col, value="Bankroll Now")
    ws.cell(row=3, column=summary_col + 1, value=br.get("balance", 500))
    real_bets = _real_bets(bets)
    ws.cell(row=4, column=summary_col, value="Total Bets")
    ws.cell(row=4, column=summary_col + 1, value=len(real_bets))
    settled = [b for b in real_bets if b.get("status") == "settled"]
    won_cnt = sum(1 for b in settled if b.get("result") == "WON")
    ws.cell(row=5, column=summary_col, value="Win Rate")
    ws.cell(row=5, column=summary_col + 1, value=f"{won_cnt}/{len(settled)}")

    # ---- SPORTS TAB ----
    ss = wb.create_sheet("Sports Bets (Manual)")

    sports_hdrs = [
        "Date", "Sport / League", "Market / Question",
        "Pinnacle Yes Odds (decimal)", "Pinnacle No Odds (decimal)",
        "De-vigged True Prob Yes", "De-vigged True Prob No",
        "Polymarket Price (ask)", "Edge (=True Prob - Ask)",
        "Stake ($)", "Shares", "Status", "Result",
        "Closing Polymarket Price", "CLV (=Closing - Ask)",
        "P&L ($)", "Notes",
    ]

    for col, hdr in enumerate(sports_hdrs, 1):
        cell = ss.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # De-vig formula in rows 2-50
    for row_n in range(2, 52):
        # Pinnacle Yes odds in col D (4), No odds in col E (5)
        d_col = get_column_letter(4)
        e_col = get_column_letter(5)
        h_col = get_column_letter(8)

        # De-vig Yes: (1/D) / (1/D + 1/E)
        ss.cell(row=row_n, column=6,
                value=f"=IF({d_col}{row_n}=\"\",\"\",ROUND((1/{d_col}{row_n})/((1/{d_col}{row_n})+(1/{e_col}{row_n})),4))")
        # De-vig No: (1/E) / (1/D + 1/E)
        ss.cell(row=row_n, column=7,
                value=f"=IF({e_col}{row_n}=\"\",\"\",ROUND((1/{e_col}{row_n})/((1/{d_col}{row_n})+(1/{e_col}{row_n})),4))")
        # Edge = de-vig yes - ask
        f_col = get_column_letter(6)
        ss.cell(row=row_n, column=9,
                value=f"=IF({f_col}{row_n}=\"\",\"\",ROUND({f_col}{row_n}-{h_col}{row_n},4))")
        # Shares = stake / ask
        j_col = get_column_letter(10)
        ss.cell(row=row_n, column=11,
                value=f"=IF({h_col}{row_n}=\"\",\"\",ROUND({j_col}{row_n}/{h_col}{row_n},4))")
        # CLV = closing - ask
        h2_col = get_column_letter(8)
        n_col = get_column_letter(14)
        ss.cell(row=row_n, column=15,
                value=f"=IF({n_col}{row_n}=\"\",\"\",ROUND({n_col}{row_n}-{h_col}{row_n},4))")

    sport_widths = [12, 16, 40, 18, 18, 18, 18, 16, 12, 10, 8, 9, 7, 20, 16, 8, 20]
    for i, w in enumerate(sport_widths, 1):
        ss.column_dimensions[get_column_letter(i)].width = w
    ss.freeze_panes = "A2"

    # Instructions row in Sports tab
    ss.cell(row=53, column=1,
            value="HOW TO USE THIS TAB: Enter Pinnacle decimal odds in columns D & E. The de-vig formula (col F) shows the true probability. Compare to Polymarket ask (col H). If Edge (col I) >= 0.08, consider placing a bet. Enter closing Polymarket price (col N) after market closes to compute CLV.").font = Font(italic=True, color="888888")

    wb.save(TRACKER_PATH)
    logger.info(f"Tracker written: {TRACKER_PATH}")
    return TRACKER_PATH


def generate_all():
    """Generate both dashboard and tracker. Returns (dashboard_path, tracker_path)."""
    d = generate_dashboard()
    t = generate_tracker()
    return d, t
